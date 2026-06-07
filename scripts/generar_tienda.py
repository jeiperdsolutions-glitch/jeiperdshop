# -*- coding: utf-8 -*-
"""
Genera la tienda HTML de JeiperdShop a partir del catalogo extraido de portatilshoprd.
- Marca: JeiperdShop (paleta Azul Tecnologico)
- Precios de venta = costo (portatilshop) + margen por tipo (margen_util.py)
- Politica de envios local RD | Pagos: transferencia / tarjeta
- HTML AUTONOMO: miniaturas base64 embebidas (se comparte 1 solo archivo)
- Interaccion: banner de ofertas, carrito persistente, favoritos, filtros,
  modo oscuro, zoom, relacionados, compartir, checkout por WhatsApp

Uso:  python generar_tienda.py
Salida: JeiperdShop.html
"""
import os
import re
import io
import glob
import json
import base64
import datetime
from openpyxl import load_workbook
from PIL import Image
from margen_util import precio_venta  # precio = costo + margen (con tope de ganancia)

CARPETA = os.path.dirname(os.path.abspath(__file__))

# ============================ CONFIG NEGOCIO ============================
CFG = {
    "marca": "Jeiperd",          # se renderiza Jeiperd + Store (Store en color acento) = JeiperdStore
    "marca2": "Store",
    "email": "jeiperdsolutions@gmail.com",
    "whatsapp": "18097925406",    # +1 809 792 5406 (confirmado por el usuario, de config.py)
    "whatsapp_show": "+1 809 792 5406",
    "instagram": "jeiperdshop",   # @jeiperdshop (sin confirmar; deja "" si no aplica)
    "pais": "Republica Dominicana",
    # URL del Apps Script (/exec) para registro/login en la Hoja de Google.
    # Mientras este vacio, el boton de Cuenta no aparece.
    "api_url": "https://script.google.com/macros/s/AKfycby0auJwANFQiJte829page21NJsE78WFYgRXPfpgsSBELi5yBYvRib4jOEZ-5kv-Qw/exec",
}

# ============================ MINIATURAS ============================
THUMB_MAX = 360
THUMB_Q = 70


def thumb_datauri(nombre_archivo):
    if not nombre_archivo:
        return ""
    ruta = os.path.join(CARPETA, "imagenes", nombre_archivo)
    if not os.path.exists(ruta):
        return ""
    try:
        im = Image.open(ruta).convert("RGB")
        im.thumbnail((THUMB_MAX, THUMB_MAX))
        buf = io.BytesIO()
        im.save(buf, format="JPEG", quality=THUMB_Q, optimize=True)
        return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception:
        return ""


# ============================ CATEGORIAS ============================
JUNK_CATS = {"todos los productos", "ver todo", "ver todos", "otros",
             "otras categorias", "otras categorías", "nuevos productos"}


def es_junk(cat):
    c = cat.strip().lower().strip("*").strip()
    return (not c) or (c in JUNK_CATS) or ("todas las ofertas" in c) or ("ofertas del mes" in c)


def limpiar_cat(cat):
    return cat.strip().strip("*").strip()


# ============================ LEER CATALOGO ============================
xlsxs = sorted(glob.glob(os.path.join(CARPETA, "portatilshop_catalogo_*.xlsx")))
xlsxs = [x for x in xlsxs if "_con_ganancia" not in os.path.basename(x)]
if not xlsxs:
    raise SystemExit("No encontre portatilshop_catalogo_*.xlsx. Corre extraer_portatilshop.py primero.")
XLSX = xlsxs[-1]
SALIDA = os.path.join(CARPETA, "JeiperdShop.html")

print(f"Leyendo {os.path.basename(XLSX)} ...")
wb = load_workbook(XLSX, read_only=True)
ws = wb.active
filas = list(ws.iter_rows(values_only=True))
encab = filas[0]
idx = {nombre: i for i, nombre in enumerate(encab)}

productos = []
cat_set = {}
suma_margen = 0.0
n_con_precio = 0
hist = {}
total_filas = len(filas) - 1
for n, f in enumerate(filas[1:], 1):
    def g(col):
        return f[idx[col]] if col in idx and idx[col] < len(f) else None

    cats = [limpiar_cat(c) for c in (g("Categorias") or "").split(",") if not es_junk(c)]
    for c in cats:
        cat_set[c] = cat_set.get(c, 0) + 1

    base = g("Precio (DOP)")
    base = float(base) if base not in (None, "") else None
    base_reg = g("Precio regular")
    base_reg = float(base_reg) if base_reg not in (None, "") else None
    on_sale = (g("En oferta") == "SI")

    nombre_p = g("Nombre") or ""
    pv = pa = None
    d = ""
    if base is not None:
        pv = precio_venta(base, cats, nombre_p)        # costo + margen, ganancia tope RD$500
        n_con_precio += 1
        mef = (pv - base) / base if base else 0
        suma_margen += mef
        hist[round(mef * 100)] = hist.get(round(mef * 100), 0) + 1
        if on_sale and base_reg and base_reg > base:
            pa = precio_venta(base_reg, cats, nombre_p)
            if pa and pv and pa > pv:
                d = str(round((1 - pv / pa) * 100)) + "%"

    if n % 100 == 0:
        print(f"  procesando {n}/{total_filas} ...")

    productos.append({
        "n": g("Nombre") or "",
        "sku": g("SKU") or "",
        "pv": pv,
        "pa": pa,
        "d": d,
        "cat": cats,
        "marca": g("Marca") or "",
        "stock": (g("Disponible") == "Si"),
        "img": thumb_datauri(g("Imagen archivo") or ""),
        "desc": (g("Descripcion corta") or "")[:400],
    })

cats_top = [c for c, _ in sorted(cat_set.items(), key=lambda x: -x[1])][:30]
total = len(productos)
ofertas = sum(1 for p in productos if p["pa"])
margen_prom = round(suma_margen / n_con_precio * 100) if n_con_precio else 0
hoy = datetime.date.today().isoformat()

print(f"{total} productos | {ofertas} en oferta | margen promedio ~{margen_prom}%")

data_json = json.dumps(productos, ensure_ascii=False)
cfg_json = json.dumps(CFG, ensure_ascii=False)
cats_json = json.dumps(cats_top, ensure_ascii=False)

# ============================ HTML ============================
HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="theme-color" content="#0A4D8C">
<title>JeiperdStore - De todo y para todos | RD</title>
<style>
  :root{
    --primario:#0A4D8C; --primario2:#0c5ba8; --acento:#00B4D8; --acento2:#02d1f7; --oscuro:#0B1F33;
    --fondo:#eef3f8; --card:#ffffff; --texto:#16263a; --gris:#6b7c8f; --borde:#e2e9f1;
    --precio:#0A4D8C; --oferta:#ff7a18; --wpp:#25D366; --fav:#e6396b;
    --sombra:0 6px 22px rgba(11,31,51,.10);
  }
  body.dark{
    --fondo:#0c1622; --card:#13212f; --texto:#e7eef6; --gris:#9fb2c2; --borde:#243443;
    --precio:#02d1f7; --sombra:0 6px 22px rgba(0,0,0,.45);
  }
  *{box-sizing:border-box;margin:0;padding:0}
  html{-webkit-text-size-adjust:100%}
  body{font-family:'Segoe UI',Roboto,Arial,sans-serif;background:var(--fondo);color:var(--texto);
       overflow-x:hidden;transition:background .25s,color .25s}
  a{text-decoration:none;color:inherit}
  img{display:block}
  .wrap{max-width:1280px;margin:auto;padding:0 18px}

  .topbar{background:var(--oscuro);color:#cfe3f5;font-size:13px}
  .topbar .wrap{display:flex;justify-content:space-between;align-items:center;padding:7px 18px;gap:10px;flex-wrap:wrap}
  .topbar b{color:var(--acento2)}.topbar a{color:#cfe3f5}

  header{background:linear-gradient(100deg,var(--primario),var(--primario2));color:#fff;
         position:sticky;top:0;z-index:60;box-shadow:0 3px 14px rgba(10,77,140,.3)}
  .hrow{display:flex;align-items:center;gap:14px;padding:14px 18px}
  .logo{font-size:27px;font-weight:800;letter-spacing:.3px;white-space:nowrap;display:flex;align-items:center;gap:8px}
  .logo .dot{width:11px;height:11px;border-radius:50%;background:var(--acento2);box-shadow:0 0 12px var(--acento2)}
  .logo span{color:var(--acento2)}
  .buscador{flex:1;display:flex;background:#fff;border-radius:10px;overflow:hidden;box-shadow:inset 0 0 0 1px rgba(0,0,0,.04)}
  .buscador input{flex:1;border:0;padding:12px 15px;font-size:16px;outline:none;color:#222}
  .buscador button{background:var(--acento);border:0;color:#04293a;font-weight:800;padding:0 22px;cursor:pointer}
  .iconbtn,.cartbtn{display:flex;align-items:center;gap:7px;font-weight:700;cursor:pointer;white-space:nowrap;
           background:rgba(255,255,255,.12);padding:10px 13px;border-radius:10px;transition:.15s;color:#fff;border:0;font-size:15px}
  .iconbtn:hover,.cartbtn:hover{background:rgba(255,255,255,.24)}
  .iconbtn.on{background:#fff;color:var(--primario)}
  .cartbtn .num,.iconbtn .num{background:var(--acento2);color:#04293a;border-radius:50%;min-width:22px;height:22px;
                display:grid;place-items:center;font-size:12px;font-weight:800;padding:0 5px}

  .catbar{background:var(--card);border-bottom:1px solid var(--borde);position:sticky;top:67px;z-index:50}
  .catbar .wrap{display:flex;gap:8px;padding:11px 18px;overflow-x:auto}
  .chip{border:1px solid var(--borde);background:var(--card);color:var(--texto);padding:7px 15px;border-radius:22px;
        font-size:13px;cursor:pointer;white-space:nowrap;transition:.15s}
  .chip:hover{border-color:var(--acento);color:var(--primario)}
  .chip.activo{background:var(--primario);color:#fff;border-color:var(--primario)}

  .hero{margin-top:18px}
  .hero-in{background:linear-gradient(115deg,var(--primario),var(--acento));color:#fff;
           border-radius:18px;padding:34px 36px;position:relative;overflow:hidden;box-shadow:var(--sombra)}
  .hero-in:after{content:"";position:absolute;right:-60px;top:-60px;width:240px;height:240px;background:rgba(255,255,255,.10);border-radius:50%}
  .hero-in h1{font-size:30px;margin-bottom:8px}.hero-in p{opacity:.95;font-size:16px;max-width:560px}
  .feats{display:flex;gap:12px;flex-wrap:wrap;margin:16px 0 2px}
  .feat{background:var(--card);border:1px solid var(--borde);border-radius:12px;padding:12px 16px;display:flex;
        align-items:center;gap:10px;font-size:14px;font-weight:600;flex:1;min-width:200px;box-shadow:var(--sombra)}
  .feat .ic{font-size:22px}

  /* banner ofertas */
  .banner{position:relative;overflow:hidden;border-radius:16px;margin-top:18px;box-shadow:var(--sombra)}
  .slides{display:flex;transition:transform .5s ease}
  .slide{min-width:100%;display:flex;gap:22px;align-items:center;padding:26px 30px;cursor:pointer;
         background:linear-gradient(110deg,#0b2c4d,#0f5ba8 70%,#00B4D8)}
  .slide img{width:170px;height:170px;object-fit:contain;background:#fff;border-radius:14px;padding:8px;flex-shrink:0}
  .sinfo{color:#fff}
  .stag{display:inline-block;background:var(--oferta);color:#fff;font-weight:800;font-size:13px;padding:4px 12px;border-radius:20px}
  .sinfo h3{font-size:22px;margin:10px 0 8px;line-height:1.25;max-width:560px}
  .sprice{font-size:24px;font-weight:800}.sprice s{font-size:16px;opacity:.7;font-weight:400;margin-left:8px}
  .slide .btn{margin-top:12px;background:#fff;color:var(--primario);width:fit-content}
  .dots{position:absolute;bottom:12px;left:50%;transform:translateX(-50%);display:flex;gap:7px}
  .dot{width:9px;height:9px;border-radius:50%;background:rgba(255,255,255,.5);cursor:pointer}
  .dot.on{background:#fff;width:22px;border-radius:5px}

  .toolbar{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;margin:22px 0 0}
  .toolbar .cnt{color:var(--gris);font-size:14px}
  .toolbar select{padding:10px 12px;border:1px solid var(--borde);border-radius:10px;background:var(--card);color:var(--texto);font-size:14px}

  /* filtros */
  .filtros{display:flex;align-items:center;gap:16px;flex-wrap:wrap;background:var(--card);border:1px solid var(--borde);
           border-radius:12px;padding:12px 16px;margin-top:12px;box-shadow:var(--sombra)}
  .fgrp{display:flex;align-items:center;gap:8px;font-size:13px;color:var(--gris);font-weight:600}
  .filtros select,.filtros input[type=number]{padding:8px 10px;border:1px solid var(--borde);border-radius:8px;
           background:var(--fondo);color:var(--texto);font-size:15px}
  .filtros input[type=number]{width:96px}
  .frange{display:flex;align-items:center;gap:6px}
  .fchk{display:flex;align-items:center;gap:6px;font-size:13px;font-weight:600;cursor:pointer;color:var(--texto)}
  .fchk input{width:17px;height:17px;accent-color:var(--primario)}
  .flimpiar{margin-left:auto;background:none;border:1px solid var(--borde);color:var(--gris);padding:8px 14px;
            border-radius:8px;cursor:pointer;font-weight:700;font-size:13px}
  .flimpiar:hover{border-color:var(--oferta);color:var(--oferta)}

  .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:18px;margin:16px 0 40px}
  .card{background:var(--card);border:1px solid var(--borde);border-radius:14px;overflow:hidden;display:flex;
        flex-direction:column;position:relative;cursor:pointer;opacity:0;transform:translateY(14px);
        transition:opacity .45s,transform .45s}
  .card.vis{opacity:1;transform:none}
  .card:hover{box-shadow:0 12px 30px rgba(10,77,140,.18);transform:translateY(-4px);border-color:var(--acento)}
  .badge{position:absolute;top:10px;left:10px;background:var(--oferta);color:#fff;font-weight:800;font-size:12px;
         padding:4px 10px;border-radius:7px;z-index:2;box-shadow:0 3px 8px rgba(255,122,24,.4)}
  .fav{position:absolute;top:9px;right:9px;z-index:3;width:34px;height:34px;border-radius:50%;cursor:pointer;
       background:rgba(255,255,255,.92);border:1px solid var(--borde);color:#cbd5e1;font-size:16px;line-height:1;
       transition:.15s}
  .fav:hover{color:var(--fav)}.fav.on{color:var(--fav);border-color:var(--fav)}
  .imgbox{height:205px;display:grid;place-items:center;padding:14px;background:#fff}
  .imgbox img{max-width:100%;max-height:100%;object-fit:contain}
  .card .body{padding:12px 14px 15px;display:flex;flex-direction:column;flex:1}
  .marca{font-size:11px;color:var(--acento);font-weight:800;text-transform:uppercase;letter-spacing:.4px;min-height:14px}
  .nombre{font-size:14px;line-height:1.35;margin:4px 0 10px;display:-webkit-box;-webkit-line-clamp:2;
          -webkit-box-orient:vertical;overflow:hidden;min-height:38px}
  .precios{margin-top:auto}
  .precio{font-size:20px;font-weight:800;color:var(--precio)}
  .pviejo{font-size:13px;color:var(--gris);text-decoration:line-through;margin-left:7px}
  .nostock{font-size:12px;color:#e05555;font-weight:700;margin-top:4px}
  .btn{margin-top:11px;background:var(--primario);color:#fff;border:0;border-radius:9px;padding:10px;font-weight:800;
       cursor:pointer;font-size:14px;transition:.15s}
  .btn:hover{background:var(--acento);color:#04293a}
  .btn:active{transform:scale(.96)}
  .masbtn{display:block;margin:6px auto 50px;background:var(--primario);color:#fff;border:0;padding:13px 44px;
          border-radius:11px;font-size:15px;font-weight:800;cursor:pointer}
  .vacio{text-align:center;color:var(--gris);padding:60px 20px;grid-column:1/-1}

  .overlay{position:fixed;inset:0;background:rgba(8,18,30,.55);display:none;z-index:90;backdrop-filter:blur(2px)}
  .overlay.open{display:block}
  .modal{position:fixed;z-index:95;top:50%;left:50%;transform:translate(-50%,-50%);width:min(860px,94vw);
         max-height:90vh;max-height:90dvh;overflow:auto;-webkit-overflow-scrolling:touch;background:var(--card);
         color:var(--texto);border-radius:16px;display:none;box-shadow:0 20px 60px rgba(0,0,0,.4)}
  .modal.open{display:block}
  .modal .x{position:absolute;top:10px;right:10px;width:38px;height:38px;font-size:24px;line-height:1;cursor:pointer;
            color:#333;background:rgba(255,255,255,.92);border:1px solid var(--borde);border-radius:50%;z-index:4}
  .mgrid{display:grid;grid-template-columns:1fr 1fr;gap:0}
  .mimg{background:#f6f9fc;display:grid;place-items:center;padding:26px;min-height:320px;overflow:hidden;position:relative}
  .mimg img{max-width:100%;max-height:340px;object-fit:contain;transition:transform .3s;cursor:zoom-in}
  .mimg img.zoom{transform:scale(1.8);cursor:zoom-out}
  .minfo{padding:26px 26px 24px}.minfo .marca{font-size:12px}
  .minfo h2{font-size:21px;line-height:1.3;margin:6px 0 12px}
  .minfo .precio{font-size:28px}
  .mdesc{color:var(--gris);font-size:14px;line-height:1.6;margin:14px 0 18px;max-height:150px;overflow:auto}
  .qty{display:flex;align-items:center;margin:12px 0 16px;width:fit-content;border:1px solid var(--borde);border-radius:10px;overflow:hidden}
  .qty button{width:40px;height:42px;border:0;background:var(--fondo);color:var(--texto);font-size:20px;cursor:pointer}
  .qty input{width:54px;height:42px;border:0;text-align:center;font-size:16px;font-weight:700;background:var(--card);color:var(--texto)}
  .mbtns{display:flex;gap:10px;flex-wrap:wrap}.mbtns .btn{flex:1;margin:0;padding:13px}
  .btn-wpp{background:var(--wpp);color:#fff}.btn-wpp:hover{background:#1fb455;color:#fff}
  .btn-sec{background:var(--fondo);color:var(--texto);border:1px solid var(--borde)}
  .btn-sec:hover{background:var(--borde);color:var(--texto)}
  .rel{padding:0 26px 24px}.rel h4{font-size:15px;margin-bottom:12px}
  .relrow{display:flex;gap:12px;overflow-x:auto;padding-bottom:6px}
  .relc{min-width:110px;width:110px;border:1px solid var(--borde);border-radius:10px;padding:8px;cursor:pointer;background:var(--card)}
  .relc:hover{border-color:var(--acento)}
  .relc img{width:100%;height:80px;object-fit:contain;background:#fff;border-radius:6px}
  .relp{font-weight:800;color:var(--precio);font-size:13px;margin-top:6px;text-align:center}

  .drawer{position:fixed;top:0;right:-440px;width:min(420px,100vw);height:100vh;height:100dvh;background:var(--card);
          color:var(--texto);z-index:100;box-shadow:-8px 0 30px rgba(0,0,0,.25);transition:right .25s;display:flex;flex-direction:column}
  .drawer.open{right:0}
  .dhead{background:var(--primario);color:#fff;padding:18px 20px;display:flex;justify-content:space-between;align-items:center}
  .dhead h3{font-size:18px}.dhead .x{background:none;border:0;color:#fff;font-size:26px;cursor:pointer}
  .ditems{flex:1;overflow:auto;-webkit-overflow-scrolling:touch;padding:14px 16px}
  .ditem{display:flex;gap:12px;padding:12px 0;border-bottom:1px solid var(--borde)}
  .ditem img{width:60px;height:60px;object-fit:contain;background:#fff;border-radius:8px}
  .ditem .it-n{font-size:13px;line-height:1.3;margin-bottom:4px}
  .ditem .it-p{font-weight:800;color:var(--precio);font-size:14px}
  .ditem .it-q{display:flex;align-items:center;gap:8px;margin-top:6px;font-size:13px}
  .ditem .it-q button{width:24px;height:24px;border:1px solid var(--borde);background:var(--fondo);color:var(--texto);border-radius:6px;cursor:pointer;font-weight:700}
  .it-del{margin-left:auto;color:#e05555;cursor:pointer;font-size:18px;background:none;border:0}
  .dempty{text-align:center;color:var(--gris);padding:50px 20px}
  .dfoot{border-top:1px solid var(--borde);padding:16px 18px;background:var(--fondo);padding-bottom:calc(16px + env(safe-area-inset-bottom))}
  .dtotal{display:flex;justify-content:space-between;font-size:18px;font-weight:800;margin-bottom:6px}
  .dnota{font-size:12px;color:var(--gris);margin-bottom:12px}
  .dfoot .btn-wpp{width:100%;border:0;border-radius:11px;padding:14px;font-size:15px;font-weight:800;cursor:pointer}

  .fab{position:fixed;right:18px;bottom:calc(18px + env(safe-area-inset-bottom));width:60px;height:60px;border-radius:50%;
       background:var(--wpp);color:#fff;display:grid;place-items:center;font-size:30px;z-index:80;
       box-shadow:0 8px 22px rgba(37,211,102,.5);cursor:pointer;text-decoration:none}
  .fab:hover{transform:scale(1.06)}

  .toast{position:fixed;left:50%;bottom:calc(22px + env(safe-area-inset-bottom));transform:translateX(-50%) translateY(20px);
         background:var(--oscuro);color:#fff;padding:12px 22px;border-radius:30px;font-weight:700;font-size:14px;z-index:120;
         opacity:0;pointer-events:none;transition:.25s;box-shadow:0 8px 24px rgba(0,0,0,.35)}
  .toast.show{opacity:1;transform:translateX(-50%) translateY(0)}

  /* registro / login */
  .authbox{position:fixed;z-index:95;top:50%;left:50%;transform:translate(-50%,-50%);width:min(380px,94vw);
           background:var(--card);color:var(--texto);border-radius:16px;display:none;padding:24px 22px 18px;
           box-shadow:0 20px 60px rgba(0,0,0,.4)}
  .authbox.open{display:block}
  .authx{position:absolute;top:10px;right:12px;background:none;border:0;font-size:26px;color:var(--gris);cursor:pointer}
  .authhead{font-size:22px;font-weight:800;display:flex;align-items:center;gap:7px;justify-content:center;margin-bottom:14px}
  .authhead .dot{width:9px;height:9px;border-radius:50%;background:var(--acento)}
  .authhead span{color:var(--acento)}
  .authtabs{display:flex;gap:8px;margin-bottom:14px}
  .authtab{flex:1;padding:10px;border:1px solid var(--borde);background:var(--fondo);color:var(--texto);
           border-radius:9px;cursor:pointer;font-weight:700;font-size:14px}
  .authtab.activo{background:var(--primario);color:#fff;border-color:var(--primario)}
  .authform{display:flex;flex-direction:column;gap:10px}
  .authform input{padding:12px;border:1px solid var(--borde);border-radius:9px;font-size:16px;background:var(--fondo);color:var(--texto)}
  .authform .btn{margin-top:4px}
  .authmsg{font-size:13px;min-height:16px;text-align:center}
  .authmsg.err{color:#e05555}.authmsg.ok{color:#16a34a}
  .authnota{font-size:11.5px;color:var(--gris);text-align:center;margin-top:12px}

  footer{background:var(--oscuro);color:#aebfd0;margin-top:20px}
  .fgrid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:26px;padding:38px 18px 8px}
  footer h4{color:#fff;font-size:15px;margin-bottom:12px}
  footer .ft-logo{font-size:23px;font-weight:800;color:#fff;margin-bottom:8px}
  footer .ft-logo span{color:var(--acento2)}
  footer p,footer li{font-size:13.5px;line-height:1.7;list-style:none}
  footer a:hover{color:var(--acento2)}
  .pay{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}
  .pay span{background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.12);border-radius:8px;padding:6px 10px;font-size:12px}
  .copy{text-align:center;border-top:1px solid rgba(255,255,255,.1);padding:16px;font-size:12.5px;opacity:.75}

  @media(max-width:760px){
    .wrap{padding:0 12px}
    .topbar{font-size:11.5px}.topbar .wrap{justify-content:center;text-align:center;gap:4px}
    .hrow{flex-wrap:wrap;gap:10px;padding:12px 12px}
    .logo{font-size:22px;flex:1}
    .buscador{order:3;flex-basis:100%}
    .iconbtn,.cartbtn{padding:9px 11px;font-size:14px}
    .catbar{position:static;top:auto}
    .hero{margin-top:12px}.hero-in{padding:22px 20px;border-radius:14px}
    .hero-in h1{font-size:22px}.hero-in p{font-size:14px}
    .feat{min-width:0;flex-basis:calc(50% - 6px);font-size:12.5px;padding:10px 12px}
    .slide{flex-direction:column;text-align:center;padding:20px;gap:12px}
    .slide img{width:130px;height:130px}.sinfo h3{font-size:17px}.sprice{font-size:20px}
    .slide .btn{margin:8px auto 0}
    .filtros{gap:10px;padding:10px 12px}.flimpiar{margin-left:0}
    .grid{grid-template-columns:repeat(2,1fr);gap:11px}
    .imgbox{height:150px}.precio{font-size:18px}.nombre{font-size:13px}.btn{padding:9px;font-size:13px}
    .mgrid{grid-template-columns:1fr}.mimg{min-height:200px;padding:18px}
    .minfo{padding:18px 18px 22px}.minfo h2{font-size:18px}.minfo .precio{font-size:24px}
    .mbtns .btn{flex-basis:100%}.rel{padding:0 18px 22px}
    .fgrid{padding:26px 14px 8px;gap:20px}
  }
  @media(max-width:380px){.grid{gap:9px}.imgbox{height:130px}}
</style>
</head>
<body>
  <div class="topbar"><div class="wrap">
    <div>&#128666; Envios a todo el pais &bull; Pago: transferencia bancaria o tarjeta</div>
    <div>&#9742; <a href="https://wa.me/__WA__" target="_blank">__WA_SHOW__</a> &bull; <a href="mailto:__EMAIL__">__EMAIL__</a></div>
  </div></div>

  <header><div class="hrow">
    <div class="logo"><span class="dot"></span>__MARCA__<span>__MARCA2__</span></div>
    <div class="buscador">
      <input id="q" type="text" placeholder="Buscar productos...">
      <button onclick="aplicar()">Buscar</button>
    </div>
    <button class="iconbtn" id="favbtn" onclick="toggleVerFavs()" title="Mis favoritos">&#10084; <span class="num" id="favnum">0</span></button>
    <button class="iconbtn" id="darkbtn" onclick="toggleDark()" title="Modo oscuro">&#127769;</button>
    <button class="iconbtn" id="cuentabtn" onclick="clicCuenta()" title="Mi cuenta" style="display:none">&#128100; <span id="cuentatxt">Cuenta</span></button>
    <div class="cartbtn" onclick="abrirCarrito()">&#128722; <span class="num" id="cartnum">0</span></div>
  </div></header>

  <div class="catbar"><div class="wrap" id="chips"></div></div>

  <div class="wrap">
    <div class="hero"><div class="hero-in">
      <h1>De todo y para todos &#128717;</h1>
      <p>Mas de __TOTAL__ productos: tecnologia, hogar, belleza, accesorios y mucho mas, con envio a todo el pais y atencion por WhatsApp.</p>
    </div></div>

    <div class="feats">
      <div class="feat"><span class="ic">&#128666;</span> Envio a todo el pais (RD$200-350)</div>
      <div class="feat"><span class="ic">&#9889;</span> Listo en 5-24 h laborables</div>
      <div class="feat"><span class="ic">&#128179;</span> Transferencia bancaria o tarjeta</div>
      <div class="feat"><span class="ic">&#128172;</span> Soporte directo por WhatsApp</div>
    </div>

    <div class="banner" id="banner"></div>

    <div class="filtros">
      <div class="fgrp">Marca
        <select id="fmarca" onchange="aplicar()"><option value="">Todas</option></select>
      </div>
      <div class="fgrp">Precio
        <div class="frange"><input id="fmin" type="number" min="0" placeholder="desde" oninput="aplicar()">
        <span>-</span><input id="fmax" type="number" min="0" placeholder="hasta" oninput="aplicar()"></div>
      </div>
      <label class="fchk"><input type="checkbox" id="fof" onchange="aplicar()"> Solo ofertas</label>
      <label class="fchk"><input type="checkbox" id="fstk" onchange="aplicar()"> Solo disponibles</label>
      <button class="flimpiar" onclick="limpiar()">Limpiar filtros</button>
    </div>

    <div class="toolbar">
      <div class="cnt" id="contador"></div>
      <select id="orden" onchange="aplicar()">
        <option value="destacados">Destacados (precios bajos primero)</option>
        <option value="pasc">Precio: menor a mayor</option>
        <option value="pdesc">Precio: mayor a menor</option>
        <option value="desc">Mayor descuento</option>
        <option value="rel">Orden original</option>
      </select>
    </div>

    <div class="grid" id="grid"></div>
    <button class="masbtn" id="mas" onclick="verMas()">Ver mas productos</button>
  </div>

  <div class="overlay" id="overlay" onclick="cerrarModal()"></div>
  <div class="modal" id="modal"></div>

  <!-- registro / login -->
  <div class="overlay" id="authoverlay" onclick="cerrarAuth()"></div>
  <div class="authbox" id="authbox">
    <button class="authx" onclick="cerrarAuth()">&times;</button>
    <div class="authhead"><span class="dot"></span>__MARCA__<span>__MARCA2__</span></div>
    <div class="authtabs">
      <button id="tab-login" class="authtab activo" onclick="authTab('login')">Iniciar sesion</button>
      <button id="tab-reg" class="authtab" onclick="authTab('reg')">Registrarse</button>
    </div>
    <div id="form-login" class="authform">
      <input id="li-correo" type="email" placeholder="Correo" autocomplete="email">
      <input id="li-clave" type="password" placeholder="Contrasena" autocomplete="current-password">
      <button class="btn" onclick="hacerLogin()">Entrar</button>
      <div class="authmsg" id="li-msg"></div>
    </div>
    <div id="form-reg" class="authform" style="display:none">
      <input id="rg-nombre" type="text" placeholder="Nombre completo">
      <input id="rg-correo" type="email" placeholder="Correo" autocomplete="email">
      <input id="rg-tel" type="tel" placeholder="Telefono / WhatsApp">
      <input id="rg-dir" type="text" placeholder="Direccion / zona de entrega">
      <input id="rg-clave" type="password" placeholder="Crea una contrasena" autocomplete="new-password">
      <button class="btn" onclick="hacerRegistro()">Crear cuenta</button>
      <div class="authmsg" id="rg-msg"></div>
    </div>
    <p class="authnota">Tus datos se guardan para agilizar tus pedidos. La contrasena se guarda cifrada.</p>
  </div>

  <div class="drawer" id="drawer">
    <div class="dhead"><h3>&#128722; Tu carrito</h3><button class="x" onclick="cerrarCarrito()">&times;</button></div>
    <div class="ditems" id="ditems"></div>
    <div class="dfoot">
      <div class="dtotal"><span>Total</span><span id="dtotal">RD$ 0</span></div>
      <div class="dnota">El envio (RD$200-350) se coordina por WhatsApp segun tu zona.</div>
      <button class="btn-wpp" onclick="checkout()">Finalizar pedido por WhatsApp</button>
    </div>
  </div>

  <a class="fab" href="https://wa.me/__WA__?text=Hola%20__MARCA____MARCA2__%2C%20quiero%20informacion" target="_blank" title="Escribenos por WhatsApp" aria-label="WhatsApp">
    <svg viewBox="0 0 32 32" width="34" height="34" fill="#fff" aria-hidden="true"><path d="M16.04 4c-6.6 0-11.96 5.36-11.96 11.96 0 2.11.55 4.17 1.6 5.99L4 28l6.21-1.63a11.93 11.93 0 0 0 5.83 1.49h.01c6.6 0 11.96-5.36 11.96-11.96S22.64 4 16.04 4zm0 21.84h-.01a9.9 9.9 0 0 1-5.05-1.38l-.36-.21-3.68.97.98-3.59-.24-.37a9.86 9.86 0 0 1-1.51-5.26c0-5.46 4.45-9.9 9.92-9.9 2.65 0 5.14 1.03 7.01 2.91a9.84 9.84 0 0 1 2.9 7c0 5.46-4.45 9.91-9.91 9.91zm5.43-7.42c-.3-.15-1.76-.87-2.03-.97-.27-.1-.47-.15-.67.15-.2.3-.77.97-.94 1.17-.17.2-.35.22-.64.07-.3-.15-1.25-.46-2.39-1.47-.88-.79-1.48-1.76-1.65-2.06-.17-.3-.02-.46.13-.61.13-.13.3-.35.45-.52.15-.17.2-.3.3-.5.1-.2.05-.37-.02-.52-.08-.15-.67-1.62-.92-2.21-.24-.58-.49-.5-.67-.51l-.57-.01c-.2 0-.52.07-.79.37-.27.3-1.04 1.02-1.04 2.48 0 1.46 1.06 2.88 1.21 3.08.15.2 2.09 3.2 5.07 4.49.71.31 1.26.49 1.69.63.71.22 1.36.19 1.87.12.57-.09 1.76-.72 2-1.41.25-.69.25-1.29.17-1.41-.07-.12-.27-.2-.57-.35z"/></svg>
  </a>
  <div class="toast" id="toast"></div>

  <footer>
    <div class="wrap fgrid">
      <div>
        <div class="ft-logo">__MARCA__<span>__MARCA2__</span></div>
        <p>Tu tienda de todo en __PAIS__. De todo y para todos, con envio a todo el pais.</p>
        <div class="pay"><span>Transferencia bancaria</span><span>Tarjeta credito/debito</span><span>WhatsApp</span></div>
      </div>
      <div>
        <h4>Politica de envios</h4>
        <ul>
          <li>&#128666; Delivery a todo el pais.</li>
          <li>&#9201; Preparacion: 5 a 24 horas laborables.</li>
          <li>&#128176; Costo: RD$200 a RD$350 segun zona.</li>
          <li>&#127991; Costo exacto se confirma por WhatsApp.</li>
          <li>&#128101; Puede recibir otra persona autorizada.</li>
        </ul>
      </div>
      <div>
        <h4>Cambios y garantia</h4>
        <ul>
          <li>Productos nuevos con garantia.</li>
          <li>Si llega defectuoso o equivocado: avisanos en 48 h con fotos y lo resolvemos.</li>
          <li>Reportes de pedido: escribe tu numero de orden y nombre.</li>
        </ul>
      </div>
      <div>
        <h4>Contacto</h4>
        <ul>
          <li>&#9742; WhatsApp: <a href="https://wa.me/__WA__" target="_blank">__WA_SHOW__</a></li>
          <li>&#9993; <a href="mailto:__EMAIL__">__EMAIL__</a></li>
          <li id="ig_li">&#128247; Instagram: <a href="https://instagram.com/__IG__" target="_blank">@__IG__</a></li>
          <li>&#128337; Respondemos en 24-48 h.</li>
        </ul>
      </div>
    </div>
    <div class="copy">&copy; __ANIO__ __MARCA____MARCA2__ &bull; __PAIS__ &bull; Catalogo actualizado __HOY__</div>
  </footer>

<script>
const CFG = __CFG__;
const PROD = __DATA__;
const CATS = __CATS__;
const PORPAG = 60;
let catActiva=null, pagina=0, filtrados=[], verFavs=false;
let cart={}, favs=new Set(), bidx=0, deals=[];

/* ---------- almacenamiento (no se pierde) ---------- */
function load(k,def){ try{ const v=localStorage.getItem(k); return v?JSON.parse(v):def; }catch(e){ return def; } }
function save(k,v){ try{ localStorage.setItem(k,JSON.stringify(v)); }catch(e){} }
function cargarEstado(){
  cart = load('jp_cart',{}) || {};
  favs = new Set(load('jp_favs',[]) || []);
  if(load('jp_dark',false)) document.body.classList.add('dark');
  sincDark();
}
function guardarCart(){ save('jp_cart',cart); }
function guardarFavs(){ save('jp_favs',[...favs]); }

function money(v){ return v==null ? 'Consultar' : 'RD$ '+Number(v).toLocaleString('es-DO'); }
function toast(msg){ const t=document.getElementById('toast'); t.textContent=msg; t.classList.add('show'); clearTimeout(t._t); t._t=setTimeout(()=>t.classList.remove('show'),1900); }

/* ---------- categorias / marcas ---------- */
function chips(){
  let h='<div class="chip activo" onclick="setCat(null,this)">Todos</div>';
  CATS.forEach(c=> h+='<div class="chip" onclick="setCat(this.dataset.c,this)" data-c="'+c.replace(/"/g,'&quot;')+'">'+c+'</div>');
  document.getElementById('chips').innerHTML=h;
}
function setCat(c,el){ catActiva=c; verFavs=false; document.getElementById('favbtn').classList.remove('on');
  document.querySelectorAll('.chip').forEach(x=>x.classList.remove('activo')); el.classList.add('activo'); aplicar(); }
function llenarMarcas(){
  const set=new Set(); PROD.forEach(p=>{ if(p.marca) set.add(p.marca); });
  const sel=document.getElementById('fmarca');
  [...set].sort((a,b)=>a.localeCompare(b)).forEach(m=>{ const o=document.createElement('option'); o.value=m; o.textContent=m; sel.appendChild(o); });
}

/* ---------- listado + filtros ---------- */
function aplicar(){
  const q=document.getElementById('q').value.toLowerCase().trim();
  const orden=document.getElementById('orden').value;
  const marca=document.getElementById('fmarca').value;
  const fmin=parseFloat(document.getElementById('fmin').value)||0;
  const fmax=parseFloat(document.getElementById('fmax').value)||Infinity;
  const soloOf=document.getElementById('fof').checked;
  const soloStk=document.getElementById('fstk').checked;
  filtrados=[];
  for(let i=0;i<PROD.length;i++){ const p=PROD[i];
    if(verFavs && !favs.has(i)) continue;
    if(catActiva && !p.cat.includes(catActiva)) continue;
    if(marca && p.marca!==marca) continue;
    if(q && !(p.n.toLowerCase().includes(q)||(p.marca||'').toLowerCase().includes(q))) continue;
    if(soloOf && !p.pa) continue;
    if(soloStk && !p.stock) continue;
    const pv=p.pv||0; if(pv<fmin || pv>fmax) continue;
    filtrados.push(i);
  }
  if(orden==='destacados') filtrados.sort((a,b)=>{
      const pa=PROD[a], pb=PROD[b];
      const sa=(pa.stock&&pa.img&&pa.pv)?0:1, sb=(pb.stock&&pb.img&&pb.pv)?0:1;
      if(sa!==sb) return sa-sb;                       // disponibles+con foto+con precio primero
      return (pa.pv||1e12)-(pb.pv||1e12);             // luego del mas barato al mas caro
    });
  else if(orden==='pasc') filtrados.sort((a,b)=>(PROD[a].pv||1e12)-(PROD[b].pv||1e12));
  else if(orden==='pdesc') filtrados.sort((a,b)=>(PROD[b].pv||0)-(PROD[a].pv||0));
  else if(orden==='desc') filtrados.sort((a,b)=>(parseInt(PROD[b].d)||0)-(parseInt(PROD[a].d)||0));
  pagina=0; document.getElementById('grid').innerHTML=''; render();
}
function limpiar(){
  document.getElementById('q').value=''; document.getElementById('fmarca').value='';
  document.getElementById('fmin').value=''; document.getElementById('fmax').value='';
  document.getElementById('fof').checked=false; document.getElementById('fstk').checked=false;
  document.getElementById('orden').value='destacados'; catActiva=null; verFavs=false;
  document.getElementById('favbtn').classList.remove('on');
  document.querySelectorAll('.chip').forEach((x,i)=>x.classList.toggle('activo',i===0));
  aplicar();
}
const io = ('IntersectionObserver' in window)
  ? new IntersectionObserver((es)=>{es.forEach(e=>{if(e.isIntersecting){e.target.classList.add('vis');io.unobserve(e.target);}})},{rootMargin:'100px'})
  : null;
function render(){
  const grid=document.getElementById('grid'), ini=pagina*PORPAG, lote=filtrados.slice(ini,ini+PORPAG);
  if(filtrados.length===0){ grid.innerHTML='<div class="vacio">'+(verFavs?'No tienes favoritos todavia. Toca el corazon en un producto.':'No se encontraron productos.')+'</div>'; }
  lote.forEach(i=>{ const p=PROD[i];
    const badge=p.d?'<div class="badge">-'+p.d+'</div>':'';
    const viejo=p.pa?'<span class="pviejo">'+money(p.pa)+'</span>':'';
    const img=p.img?'<img loading="lazy" src="'+p.img+'">':'<div style="color:#bbb">sin foto</div>';
    const stock=p.stock?'':'<div class="nostock">Agotado</div>';
    const d=document.createElement('div');
    d.className='card'; d.onclick=()=>verProducto(i);
    d.innerHTML=badge+'<button class="fav'+(favs.has(i)?' on':'')+'" onclick="event.stopPropagation();toggleFav('+i+',this)">&#10084;</button>'+
      '<div class="imgbox">'+img+'</div><div class="body">'+
      '<div class="marca">'+(p.marca||'')+'</div><div class="nombre">'+p.n+'</div>'+
      '<div class="precios"><span class="precio">'+money(p.pv)+'</span>'+viejo+stock+'</div>'+
      '<button class="btn" onclick="event.stopPropagation();addCart('+i+',1)">Agregar al carrito</button></div>';
    grid.appendChild(d); if(io) io.observe(d); else d.classList.add('vis');
  });
  document.getElementById('contador').textContent=filtrados.length+' productos';
  document.getElementById('mas').style.display=(ini+PORPAG)<filtrados.length?'block':'none';
}
function verMas(){ pagina++; render(); }

/* ---------- modal producto ---------- */
function verProducto(i){
  const p=PROD[i];
  const badge=p.d?'<div class="badge">-'+p.d+'</div>':'';
  const viejo=p.pa?'<span class="pviejo">'+money(p.pa)+'</span>':'';
  const img=p.img?'<img onclick="this.classList.toggle(\'zoom\')" src="'+p.img+'">':'<div style="color:#bbb">sin foto</div>';
  const desc=p.desc?'<div class="mdesc">'+p.desc+'</div>':'';
  const stock=p.stock?'':'<div class="nostock">Producto agotado</div>';
  document.getElementById('modal').innerHTML=
    '<button class="x" onclick="cerrarModal()">&times;</button>'+
    '<div class="mgrid"><div class="mimg">'+badge+img+'</div><div class="minfo">'+
      '<div class="marca">'+(p.marca||'')+'</div><h2>'+p.n+'</h2>'+
      '<div class="precios"><span class="precio">'+money(p.pv)+'</span>'+viejo+'</div>'+stock+desc+
      '<div class="qty"><button onclick="chgQ(-1)">-</button><input id="mq" value="1" readonly><button onclick="chgQ(1)">+</button></div>'+
      '<div class="mbtns">'+
        '<button class="btn" onclick="addCart('+i+',+document.getElementById(\'mq\').value);cerrarModal();abrirCarrito()">Agregar al carrito</button>'+
        '<button class="btn btn-wpp" onclick="comprarYa('+i+')">Comprar por WhatsApp</button>'+
        '<button class="btn btn-sec" onclick="compartir('+i+')">&#128279; Compartir</button>'+
      '</div></div></div>'+ relacionados(i);
  document.getElementById('overlay').classList.add('open');
  document.getElementById('modal').classList.add('open');
  document.getElementById('modal').scrollTop=0;
}
function relacionados(i){
  const p=PROD[i], out=[];
  for(let j=0;j<PROD.length && out.length<8;j++){ if(j===i) continue; const q=PROD[j];
    if(q.img && q.cat.some(c=>p.cat.includes(c))) out.push(j); }
  if(!out.length) return '';
  return '<div class="rel"><h4>Productos relacionados</h4><div class="relrow">'+
    out.map(j=>'<div class="relc" onclick="verProducto('+j+')"><img src="'+PROD[j].img+'"><div class="relp">'+money(PROD[j].pv)+'</div></div>').join('')+'</div></div>';
}
function chgQ(n){ const e=document.getElementById('mq'); e.value=Math.max(1,(+e.value)+n); }
function cerrarModal(){ document.getElementById('overlay').classList.remove('open'); document.getElementById('modal').classList.remove('open'); }

/* ---------- favoritos ---------- */
function toggleFav(i,el){
  if(favs.has(i)){ favs.delete(i); if(el)el.classList.remove('on'); toast('Quitado de favoritos'); }
  else{ favs.add(i); if(el)el.classList.add('on'); toast('❤ Agregado a favoritos'); }
  guardarFavs(); document.getElementById('favnum').textContent=favs.size;
  if(verFavs) aplicar();
}
function toggleVerFavs(){ verFavs=!verFavs; document.getElementById('favbtn').classList.toggle('on',verFavs);
  if(verFavs){ catActiva=null; document.querySelectorAll('.chip').forEach((x,i)=>x.classList.toggle('activo',i===0)); } aplicar(); }

/* ---------- carrito ---------- */
function addCart(i,q){ cart[i]=(cart[i]||0)+(q||1); guardarCart(); refrescaCarrito(); flash(); toast('✓ Agregado al carrito'); }
function quita(i){ if(cart[i]){ cart[i]--; if(cart[i]<=0) delete cart[i]; } guardarCart(); refrescaCarrito(); }
function sumaCart(i){ cart[i]++; guardarCart(); refrescaCarrito(); }
function borra(i){ delete cart[i]; guardarCart(); refrescaCarrito(); }
function totalCart(){ let t=0; for(const i in cart){ t+=(PROD[i].pv||0)*cart[i]; } return t; }
function nItems(){ let n=0; for(const i in cart) n+=cart[i]; return n; }
function refrescaCarrito(){
  document.getElementById('cartnum').textContent=nItems();
  const cont=document.getElementById('ditems'), keys=Object.keys(cart);
  if(keys.length===0){ cont.innerHTML='<div class="dempty">Tu carrito esta vacio.<br>Agrega productos para pedir por WhatsApp.</div>'; }
  else{ cont.innerHTML=keys.map(i=>{ const p=PROD[i]; return '<div class="ditem">'+
      (p.img?'<img src="'+p.img+'">':'')+'<div style="flex:1"><div class="it-n">'+p.n+'</div>'+
      '<div class="it-p">'+money(p.pv)+'</div><div class="it-q">'+
      '<button onclick="quita('+i+')">-</button>'+cart[i]+'<button onclick="sumaCart('+i+')">+</button>'+
      '<button class="it-del" onclick="borra('+i+')" title="Quitar">&times;</button></div></div></div>'; }).join(''); }
  document.getElementById('dtotal').textContent=money(totalCart());
}
function abrirCarrito(){ document.getElementById('drawer').classList.add('open'); }
function cerrarCarrito(){ document.getElementById('drawer').classList.remove('open'); }
function flash(){ const b=document.getElementById('cartnum'); b.style.transform='scale(1.5)'; setTimeout(()=>b.style.transform='',180); }

/* ---------- WhatsApp ---------- */
function comprarYa(i){ const q=+document.getElementById('mq').value, p=PROD[i];
  const msg='Hola '+CFG.marca+CFG.marca2+', quiero comprar:%0A%0A- '+p.n+' (x'+q+') = '+money((p.pv||0)*q)+'%0A%0ATotal: '+money((p.pv||0)*q)+'%0A%0AMi zona de entrega es: ';
  window.open('https://wa.me/'+CFG.whatsapp+'?text='+msg,'_blank'); }
/* ---- cotizacion con fotos (imagen) ---- */
function _loadImg(src){ return new Promise(r=>{ const im=new Image(); im.onload=()=>r(im); im.onerror=()=>r(null); im.src=src; }); }
function _contain(c,im,x,y,w,h){ const r=Math.min(w/im.width,h/im.height); const dw=im.width*r,dh=im.height*r; c.drawImage(im,x+(w-dw)/2,y+(h-dh)/2,dw,dh); }
function _wrap(c,txt,x,y,maxw,lh,maxlines){ const ws=txt.split(' '); let line='',ln=0; for(let k=0;k<ws.length;k++){ const t=line?line+' '+ws[k]:ws[k]; if(c.measureText(t).width>maxw && line){ c.fillText(line,x,y); y+=lh; line=ws[k]; if(++ln>=maxlines-1){ let rest=ws.slice(k).join(' '); while(c.measureText(rest+'...').width>maxw && rest.length>3) rest=rest.slice(0,-1); c.fillText(rest+'...',x,y); return; } } else line=t; } c.fillText(line,x,y); }

async function cotizacionCanvas(){
  const keys=Object.keys(cart), imgs={};
  await Promise.all(keys.map(async i=>{ if(PROD[i].img) imgs[i]=await _loadImg(PROD[i].img); }));
  const top=160, rowH=80, foot=120;
  const cv=document.createElement('canvas'); cv.width=820; cv.height=top+keys.length*rowH+foot;
  const c=cv.getContext('2d');
  c.fillStyle='#fff'; c.fillRect(0,0,cv.width,cv.height);
  c.fillStyle='#0B1F33'; c.fillRect(0,0,cv.width,112);
  c.textBaseline='alphabetic';
  c.font='bold 34px Arial'; c.fillStyle='#fff'; c.fillText('Jeiperd',30,54);
  const wj=c.measureText('Jeiperd').width; c.fillStyle='#02d1f7'; c.fillText('Store',30+wj,54);
  c.textAlign='right'; c.font='bold 22px Arial'; c.fillStyle='#fff'; c.fillText('COTIZACION',792,46); c.textAlign='left';
  const fecha=new Date().toLocaleDateString('es-DO'); const folio='JS-'+String(Date.now()).slice(-6);
  c.font='13px Arial'; c.fillStyle='#cfe3f5'; c.fillText('Fecha: '+fecha+'     Folio: '+folio,30,90);
  c.fillStyle='#16263a'; c.font='13px Arial';
  c.fillText(usuario?('Cliente: '+(usuario.nombre||'')+'   Tel: '+(usuario.telefono||'')+'   '+(usuario.direccion||'')):'Cliente: (se completa al confirmar por WhatsApp)',30,140);
  c.textBaseline='middle';
  keys.forEach((i,idx)=>{ const p=PROD[i], ry=top+idx*rowH;
    c.strokeStyle='#e2e9f1'; c.beginPath(); c.moveTo(20,ry); c.lineTo(800,ry); c.stroke();
    c.fillStyle='#f6f9fc'; c.fillRect(26,ry+10,60,60); if(imgs[i]) _contain(c,imgs[i],26,ry+10,60,60);
    c.fillStyle='#16263a'; c.font='14px Arial'; _wrap(c,p.n,102,ry+28,470,18,2);
    c.fillStyle='#6b7c8f'; c.font='13px Arial'; c.fillText('Cantidad: '+cart[i]+'   x   '+money(p.pv),102,ry+60);
    c.fillStyle='#0A4D8C'; c.font='bold 15px Arial'; c.textAlign='right'; c.fillText(money((p.pv||0)*cart[i]),794,ry+40); c.textAlign='left';
  });
  const ty=top+keys.length*rowH+8;
  c.strokeStyle='#0B1F33'; c.lineWidth=2; c.beginPath(); c.moveTo(20,ty); c.lineTo(800,ty); c.stroke(); c.lineWidth=1;
  c.fillStyle='#16263a'; c.font='bold 20px Arial'; c.fillText('TOTAL',30,ty+34);
  c.textAlign='right'; c.fillStyle='#0A4D8C'; c.font='bold 24px Arial'; c.fillText(money(totalCart()),794,ty+36); c.textAlign='left';
  c.fillStyle='#6b7c8f'; c.font='12px Arial'; c.textBaseline='alphabetic'; c.fillText('* El envio (RD$200-350) se suma segun la zona. Pago: transferencia o tarjeta.',30,ty+62);
  c.fillStyle='#0B1F33'; c.fillRect(0,cv.height-42,cv.width,42);
  c.fillStyle='#fff'; c.font='13px Arial'; c.textAlign='center';
  c.fillText('JeiperdStore  -  WhatsApp '+CFG.whatsapp_show+'  -  De todo y para todos',cv.width/2,cv.height-16); c.textAlign='left';
  return cv;
}

function _msgPedido(plano){
  const nl = plano?'\n':'%0A'; const keys=Object.keys(cart);
  let m='Hola '+CFG.marca+CFG.marca2+', quiero hacer este pedido:'+nl+nl;
  keys.forEach(i=>{ const p=PROD[i]; m+='- '+p.n+' (x'+cart[i]+') = '+money((p.pv||0)*cart[i])+nl; });
  m+=nl+'Total productos: '+money(totalCart())+nl+'(El envio se suma segun mi zona)'+nl+nl;
  m+='Mi nombre: '+(usuario?usuario.nombre||'':'')+nl+'Mi telefono: '+(usuario?usuario.telefono||'':'')+nl+'Mi direccion/zona: '+(usuario?usuario.direccion||'':'');
  return plano?m:m.split(' ').join('%20');
}

async function checkout(){
  const keys=Object.keys(cart);
  if(keys.length===0){ toast('Tu carrito esta vacio'); return; }
  toast('Preparando tu cotizacion...');
  const textoURL=encodeURIComponent(_msgPedido(true));
  const waURL='https://wa.me/'+CFG.whatsapp+'?text='+textoURL;
  let cv=null; try{ cv=await cotizacionCanvas(); }catch(e){ cv=null; }
  if(!cv){ window.open(waURL,'_blank'); return; }
  cv.toBlob(async (blob)=>{
    const file=new File([blob],'cotizacion-jeiperdstore.png',{type:'image/png'});
    if(navigator.canShare && navigator.canShare({files:[file]})){
      try{ await navigator.share({files:[file], text:_msgPedido(true)}); return; }catch(e){}
    }
    const a=document.createElement('a'); a.href=URL.createObjectURL(blob); a.download='cotizacion-jeiperdstore.png'; document.body.appendChild(a); a.click(); a.remove();
    toast('Cotizacion descargada — adjuntala en WhatsApp');
    setTimeout(()=>window.open(waURL,'_blank'),900);
  },'image/png');
}
function compartir(i){ const p=PROD[i], txt=p.n+' - '+money(p.pv)+' en '+CFG.marca+CFG.marca2;
  if(navigator.share){ navigator.share({title:CFG.marca+CFG.marca2,text:txt}).catch(()=>{}); }
  else{ window.open('https://wa.me/?text='+encodeURIComponent(txt),'_blank'); } }

/* ---------- banner ofertas ---------- */
function initBanner(){
  deals=[]; for(let i=0;i<PROD.length;i++){ if(PROD[i].pa && PROD[i].img) deals.push(i); }
  deals.sort((a,b)=>(parseInt(PROD[b].d)||0)-(parseInt(PROD[a].d)||0)); deals=deals.slice(0,8);
  if(!deals.length){ document.getElementById('banner').style.display='none'; return; }
  const slides=deals.map(i=>{ const p=PROD[i]; return '<div class="slide" onclick="verProducto('+i+')">'+
    '<img src="'+p.img+'"><div class="sinfo"><span class="stag">OFERTA -'+p.d+'</span>'+
    '<h3>'+p.n+'</h3><div class="sprice">'+money(p.pv)+' <s>'+money(p.pa)+'</s></div>'+
    '<button class="btn">Ver oferta</button></div></div>'; }).join('');
  const dots=deals.map((_,k)=>'<span class="dot'+(k===0?' on':'')+'" onclick="event.stopPropagation();goSlide('+k+')"></span>').join('');
  document.getElementById('banner').innerHTML='<div class="slides" id="slides">'+slides+'</div><div class="dots">'+dots+'</div>';
  setInterval(()=>goSlide((bidx+1)%deals.length),4500);
}
function goSlide(k){ bidx=k; document.getElementById('slides').style.transform='translateX(-'+(k*100)+'%)';
  document.querySelectorAll('.banner .dot').forEach((d,j)=>d.classList.toggle('on',j===k)); }

/* ---------- modo oscuro ---------- */
function sincDark(){ document.getElementById('darkbtn').innerHTML = document.body.classList.contains('dark') ? '&#9728;' : '&#127769;'; }
function toggleDark(){ document.body.classList.toggle('dark'); save('jp_dark',document.body.classList.contains('dark')); sincDark(); }

/* ---------- cuenta: registro / login (Hoja de Google via Apps Script) ---------- */
let usuario = null;
async function sha256(txt){
  const buf = await crypto.subtle.digest('SHA-256', new TextEncoder().encode(txt));
  return [...new Uint8Array(buf)].map(b=>b.toString(16).padStart(2,'0')).join('');
}
function initCuenta(){
  if(!CFG.api_url) return;                       // sin backend configurado, no mostrar
  document.getElementById('cuentabtn').style.display='';
  usuario = load('jp_user', null);
  pintarCuenta();
}
function pintarCuenta(){
  document.getElementById('cuentatxt').textContent = usuario ? (String(usuario.nombre||'Cuenta').split(' ')[0]) : 'Cuenta';
}
function clicCuenta(){ usuario ? menuCuenta() : abrirAuth(); }
function abrirAuth(){ document.getElementById('authoverlay').classList.add('open'); document.getElementById('authbox').classList.add('open'); }
function cerrarAuth(){ document.getElementById('authoverlay').classList.remove('open'); document.getElementById('authbox').classList.remove('open'); }
function authTab(cual){
  document.getElementById('tab-login').classList.toggle('activo',cual==='login');
  document.getElementById('tab-reg').classList.toggle('activo',cual==='reg');
  document.getElementById('form-login').style.display = cual==='login'?'flex':'none';
  document.getElementById('form-reg').style.display = cual==='reg'?'flex':'none';
}
function menuCuenta(){
  if(confirm('Sesion de '+(usuario.nombre||usuario.correo)+'.\n\nAceptar = cerrar sesion.')){
    usuario=null; save('jp_user',null); pintarCuenta(); toast('Sesion cerrada');
  }
}
function jsonp(url){
  return new Promise((resolve,reject)=>{
    const cb='jpcb_'+Math.random().toString(36).slice(2);
    const s=document.createElement('script');
    window[cb]=(data)=>{ resolve(data); try{delete window[cb];}catch(e){} s.remove(); };
    s.onerror=()=>{ reject(new Error('red')); s.remove(); };
    s.src=url+(url.includes('?')?'&':'?')+'callback='+cb;
    document.body.appendChild(s);
    setTimeout(()=>{ if(window[cb]){ try{delete window[cb];}catch(e){} s.remove(); reject(new Error('timeout')); } },12000);
  });
}
async function hacerRegistro(){
  const m=document.getElementById('rg-msg'); m.className='authmsg'; m.textContent='Guardando...';
  const nombre=document.getElementById('rg-nombre').value.trim();
  const correo=document.getElementById('rg-correo').value.trim();
  const tel=document.getElementById('rg-tel').value.trim();
  const dir=document.getElementById('rg-dir').value.trim();
  const clave=document.getElementById('rg-clave').value;
  if(!nombre||!correo||!clave){ m.className='authmsg err'; m.textContent='Completa nombre, correo y contrasena.'; return; }
  const hash=await sha256(clave);
  try{
    await fetch(CFG.api_url,{method:'POST',mode:'no-cors',
      headers:{'Content-Type':'text/plain;charset=utf-8'},
      body:JSON.stringify({nombre,correo,telefono:tel,direccion:dir,clave:hash})});
    usuario={nombre,correo,telefono:tel,direccion:dir}; save('jp_user',usuario); pintarCuenta();
    m.className='authmsg ok'; m.textContent='Cuenta creada. Bienvenido, '+nombre.split(' ')[0]+'!';
    toast('Cuenta creada'); setTimeout(cerrarAuth,1300);
  }catch(e){ m.className='authmsg err'; m.textContent='No se pudo registrar. Intenta luego.'; }
}
async function hacerLogin(){
  const m=document.getElementById('li-msg'); m.className='authmsg'; m.textContent='Entrando...';
  const correo=document.getElementById('li-correo').value.trim();
  const clave=document.getElementById('li-clave').value;
  if(!correo||!clave){ m.className='authmsg err'; m.textContent='Escribe correo y contrasena.'; return; }
  const hash=await sha256(clave);
  try{
    const r=await jsonp(CFG.api_url+'?action=login&correo='+encodeURIComponent(correo)+'&clave='+hash);
    if(r&&r.ok){ usuario=r.user; save('jp_user',usuario); pintarCuenta();
      m.className='authmsg ok'; m.textContent='Hola, '+String(usuario.nombre||'').split(' ')[0]+'!';
      toast('Sesion iniciada'); setTimeout(cerrarAuth,1000); }
    else { m.className='authmsg err'; m.textContent=(r&&r.error)||'No se pudo entrar.'; }
  }catch(e){ m.className='authmsg err'; m.textContent='No se pudo conectar al servidor.'; }
}

/* ---------- init ---------- */
document.getElementById('q').addEventListener('keydown',e=>{ if(e.key==='Enter') aplicar(); });
document.addEventListener('keydown',e=>{ if(e.key==='Escape'){ cerrarModal(); cerrarCarrito(); cerrarAuth(); } });
if(!CFG.instagram){ document.getElementById('ig_li').style.display='none'; }
cargarEstado(); chips(); llenarMarcas(); initBanner(); aplicar(); refrescaCarrito(); initCuenta();
document.getElementById('favnum').textContent=favs.size;
</script>
</body>
</html>"""

rep = {
    "__DATA__": data_json, "__CFG__": cfg_json, "__CATS__": cats_json,
    "__TOTAL__": str(total), "__HOY__": hoy, "__ANIO__": hoy[:4],
    "__MARCA__": CFG["marca"], "__MARCA2__": CFG["marca2"],
    "__EMAIL__": CFG["email"], "__WA__": CFG["whatsapp"],
    "__WA_SHOW__": CFG["whatsapp_show"], "__IG__": CFG["instagram"],
    "__PAIS__": CFG["pais"],
}
for k, v in rep.items():
    HTML = HTML.replace(k, v)

with open(SALIDA, "w", encoding="utf-8") as fh:
    fh.write(HTML)

print(f"LISTO -> {SALIDA}")
print(f"Margen promedio aplicado: ~{margen_prom}%")
