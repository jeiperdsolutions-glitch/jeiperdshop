# -*- coding: utf-8 -*-
"""
Extractor de catalogo completo de portatilshoprd.com (WooCommerce Store API).
Genera un Excel con todos los productos y descarga la imagen principal de cada uno.

Uso:
    python extraer_portatilshop.py

Salida (dentro de la misma carpeta del script):
    portatilshop_catalogo_<fecha>.xlsx
    imagenes/<SKU>.jpg
"""
import os
import re
import time
import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = "https://portatilshoprd.com/wp-json/wc/store/v1/products"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")
HEADERS = {
    "User-Agent": UA,
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "es-DO,es;q=0.9,en;q=0.8",
    "Referer": "https://portatilshoprd.com/tienda/",
    "Origin": "https://portatilshoprd.com",
    "Connection": "keep-alive",
    "sec-ch-ua": '"Chromium";v="123", "Not:A-Brand";v="99"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
}

CARPETA = os.path.dirname(os.path.abspath(__file__))
HOY = datetime.date.today().isoformat()
XLSX = os.path.join(CARPETA, f"portatilshop_catalogo_{HOY}.xlsx")
IMG_DIR = os.path.join(CARPETA, "imagenes")
# OJO: Windows no distingue mayus/minus, asi que NO usar "bitacora.md"
# (chocaria con BITACORA.md). Este extractor escribe en su propio log.
BITACORA = os.path.join(CARPETA, "extraccion_log.md")

DESCARGAR_IMAGENES = True   # ponlo en False si solo quieres el Excel
PER_PAGE = 100


def limpiar_html(texto):
    if not texto:
        return ""
    texto = re.sub(r"<[^>]+>", " ", texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def precio(prod, campo):
    """Convierte el precio en unidades menores (centavos) a numero decimal."""
    prices = prod.get("prices") or {}
    val = prices.get(campo)
    if val in (None, ""):
        return None
    try:
        minor = int(prices.get("currency_minor_unit", 2))
        return round(int(val) / (10 ** minor), 2)
    except (ValueError, TypeError):
        return None


def nombre_archivo_seguro(s):
    return re.sub(r"[^A-Za-z0-9_.-]", "_", str(s))[:80]


def _cargar_cookies_cf(sesion):
    """Si existe cf_cookies.json (creado por solve_cloudflare.py con Playwright),
    usa esas cookies + User-Agent para pasar el reto de Cloudflare (entorno nube)."""
    import json
    ruta = os.path.join(CARPETA, "cf_cookies.json")
    if not os.path.exists(ruta):
        return False
    try:
        data = json.load(open(ruta, encoding="utf-8"))
        if data.get("ua"):
            sesion.headers["User-Agent"] = data["ua"]
        n = 0
        for c in data.get("cookies", []):
            sesion.cookies.set(c["name"], c["value"],
                               domain=c.get("domain", "portatilshoprd.com"),
                               path=c.get("path", "/"))
            n += 1
        print(f"  (Cloudflare: cargadas {n} cookies del navegador)")
        return True
    except Exception as e:
        print(f"  (no pude cargar cf_cookies.json: {e})")
        return False


def bajar_todos():
    sesion = requests.Session()
    sesion.headers.update(HEADERS)
    uso_cf = _cargar_cookies_cf(sesion)
    # Si no hay cookies de navegador, "primar" visitando la web (sirve desde IP residencial)
    if not uso_cf:
        try:
            sesion.get("https://portatilshoprd.com/", timeout=30)
            time.sleep(1)
        except Exception:
            pass
    productos = []
    pagina = 1
    while True:
        r = sesion.get(BASE, params={"per_page": PER_PAGE, "page": pagina}, timeout=30)
        if r.status_code != 200:
            print(f"  ! pagina {pagina} devolvio {r.status_code}, reintentando...")
            time.sleep(5)
            r = sesion.get(BASE, params={"per_page": PER_PAGE, "page": pagina}, timeout=30)
        if r.status_code != 200:
            raise SystemExit(f"Cloudflare bloqueo la peticion (HTTP {r.status_code}). "
                             f"Primeros 200 chars: {r.text[:200]!r}")
        lote = r.json()
        if not lote:
            break
        productos.extend(lote)
        total_paginas = r.headers.get("X-WP-TotalPages", "?")
        print(f"  pagina {pagina}/{total_paginas} -> {len(lote)} productos (acumulado {len(productos)})")
        if pagina >= int(r.headers.get("X-WP-TotalPages", pagina)):
            break
        pagina += 1
        time.sleep(0.5)  # cortesia con el servidor
    return productos, sesion


def construir_excel(productos, sesion):
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogo"
    cols = ["ID", "SKU", "Nombre", "Precio (DOP)", "Precio regular", "En oferta",
            "% Desc", "Stock", "Disponible", "Categorias", "Marca",
            "Descripcion corta", "Link", "Imagen URL", "Imagen archivo"]
    ws.append(cols)

    # estilo encabezado
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(vertical="center")

    if DESCARGAR_IMAGENES:
        os.makedirs(IMG_DIR, exist_ok=True)

    en_oferta_count = 0
    img_ok = 0
    for i, p in enumerate(productos, 1):
        p_actual = precio(p, "price")
        p_regular = precio(p, "regular_price")
        on_sale = bool(p.get("on_sale"))
        if on_sale:
            en_oferta_count += 1
        desc = None
        if on_sale and p_regular and p_actual and p_regular > 0:
            desc = round((1 - p_actual / p_regular) * 100)

        sku = p.get("sku") or f"id{p.get('id')}"
        imgs = p.get("images") or []
        img_url = imgs[0].get("src") if imgs else ""
        img_file = ""

        if DESCARGAR_IMAGENES and img_url:
            ext = os.path.splitext(img_url.split("?")[0])[1] or ".jpg"
            nombre = nombre_archivo_seguro(sku) + ext
            ruta = os.path.join(IMG_DIR, nombre)
            if not os.path.exists(ruta):
                try:
                    ir = sesion.get(img_url, timeout=30)
                    if ir.status_code == 200:
                        with open(ruta, "wb") as f:
                            f.write(ir.content)
                        img_file = nombre
                        img_ok += 1
                except Exception as e:
                    print(f"    ! imagen fallo {sku}: {e}")
            else:
                img_file = nombre
                img_ok += 1

        cats = ", ".join(c.get("name", "") for c in (p.get("categories") or []))
        marcas = ", ".join(b.get("name", "") for b in (p.get("brands") or []))
        stock = (p.get("stock_availability") or {}).get("class", "")

        ws.append([
            p.get("id"), sku, p.get("name", ""), p_actual, p_regular,
            "SI" if on_sale else "", (f"{desc}%" if desc is not None else ""),
            stock, "Si" if p.get("is_in_stock") else "No",
            cats, marcas, limpiar_html(p.get("short_description")),
            p.get("permalink", ""), img_url, img_file,
        ])

        if i % 100 == 0:
            print(f"  procesados {i}/{len(productos)} (imagenes ok: {img_ok})")

    # resaltar filas en oferta
    naranja = PatternFill("solid", fgColor="FFF2CC")
    for fila in ws.iter_rows(min_row=2):
        if fila[5].value == "SI":  # columna "En oferta"
            for celda in fila:
                celda.fill = naranja

    # anchos
    anchos = {"A": 8, "B": 14, "C": 55, "D": 14, "E": 14, "F": 9, "G": 8,
              "H": 12, "I": 10, "J": 40, "K": 16, "L": 50, "M": 45, "N": 45, "O": 18}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    wb.save(XLSX)
    return en_oferta_count, img_ok


def escribir_bitacora(total, ofertas, img_ok):
    linea = (f"\n## {HOY} - Extraccion portatilshoprd.com\n"
             f"- Productos extraidos: **{total}**\n"
             f"- En oferta: {ofertas}\n"
             f"- Imagenes descargadas: {img_ok}\n"
             f"- Excel: `{os.path.basename(XLSX)}`\n"
             f"- Fuente: WooCommerce Store API (wc/store/v1/products)\n")
    modo = "a" if os.path.exists(BITACORA) else "w"
    with open(BITACORA, modo, encoding="utf-8") as f:
        if modo == "w":
            f.write("# Log de extracciones (portatilshop)\n")
        f.write(linea)


def main():
    print("Descargando catalogo de portatilshoprd.com ...")
    productos, sesion = bajar_todos()
    print(f"Total productos: {len(productos)}")
    print("Generando Excel y descargando imagenes ...")
    ofertas, img_ok = construir_excel(productos, sesion)
    escribir_bitacora(len(productos), ofertas, img_ok)
    print("\n=== LISTO ===")
    print(f"Excel:     {XLSX}")
    print(f"Imagenes:  {IMG_DIR}  ({img_ok} archivos)")
    print(f"En oferta: {ofertas}")
    print(f"Bitacora:  {BITACORA}")


if __name__ == "__main__":
    main()
